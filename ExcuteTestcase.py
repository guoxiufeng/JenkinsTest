# encoding:utf-8

import Key_Words
import openpyxl
import os
import Gloabl_var
import Common
import datetime

dict_testdata = {}


def key_words(action_name, action_params):

        if action_name == 'Login':
            parms = action_params.split(",")
            Key_Words.login(dict_testdata[parms[0]], dict_testdata[parms[1]])
        if action_name == 'LoginOut':
            Key_Words.log_out()


def get_data(filename, sheetname):

    dir_case = './AutoTest/' + filename + '.xlsx'
    wb = openpyxl.load_workbook(dir_case)
    sheet = wb[sheetname]
    dict={}
    for i in range(1, sheet.max_row):

        for j in range(1, sheet.max_column + 1):
            key = sheet.cell(row=1, column=j).value
            val = sheet.cell(row=2, column=j).value
            dict[key] = val
    wb.close()
    return dict


def start_execute(filename, sheetname):
    dir_case = './AutoTest/Test_Script/' + filename + '.xlsx'
    wb = openpyxl.load_workbook(dir_case)
    sheet = wb[sheetname]
    body = []
    for i in range(1, sheet.max_row):
        desc = sheet.cell(row=i+1, column=1).value
        action_name = sheet.cell(row=i+1, column=2).value
        action_params = sheet.cell(row=i+1, column=3).value
        exception_desc = ''
        try:
            key_words(action_name, action_params)
        except Exception as ex_results:
            exception_desc = ex_results
            print(ex_results)
        finally:
            execute_res = ''
            if exception_desc != '':
                execute_res = 'Fail'
            else:
                execute_res = 'Pass'

            result = {'ID': i, 'Desc': desc, 'Result': execute_res}
            body.append(result)

    dirs = os.listdir(Gloabl_var.picture_path)
    dirs = sorted(dirs, key = lambda x: int(x[:-4]))
    Common.create_gif(dirs, Gloabl_var.picture_path + "/result.gif", 1)
    Common.generate_html(body, Gloabl_var.starttime, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    wb.close()


def execute_testcase(filename, sheetname):
    dir_case = './AutoTest/' + filename + '.xlsx'
    wb = openpyxl.load_workbook(dir_case)
    # �ӱ��л�ȡ��Ԫ�������
    # ��ǰ��Ծ�ı�
    sheet = wb[sheetname]
    for i in range(1, sheet.max_row):
        execute_flag = sheet.cell(row = i + 1, column = 1).value
        execute_testcasename = sheet.cell(row = i + 1, column = 2).value
        testcase_datasheetname = sheet.cell(row = i + 1, column = 2).value
        testcase_sheetname = sheet.cell(row = i + 1, column = 2).value

        if execute_flag == 'Y':
            global dict_testdata
            dict_testdata = get_data('testdata', testcase_datasheetname)
            start_execute(execute_testcasename, testcase_sheetname)
    wb.close()


if __name__ == '__main__':
    execute_testcase('configure', 'configure')


